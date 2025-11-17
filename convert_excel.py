#!/usr/bin/env python3.11
"""
Excel to JSON Converter for S4 League Player Statistics

This script converts the S4LeagueAutomatisierung.xlsx file to JSON format
for use in the web dashboard.

Usage:
    python3.11 convert_excel.py

Requirements:
    pip install openpyxl
"""

import openpyxl
import json
import sys
from pathlib import Path

def convert_excel_to_json(excel_path: str, output_path: str) -> None:
    """
    Convert Excel file to JSON format.
    
    Args:
        excel_path: Path to the Excel file
        output_path: Path where the JSON file should be saved
    """
    print(f"Loading Excel file: {excel_path}")
    
    # Load workbook with data_only=True to get calculated values instead of formulas
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Convert all sheets to JSON
    data = {}
    
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        sheet_data = []
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Process data rows (starting from row 2)
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    # Only add non-None values
                    if value is not None:
                        row_dict[headers[i]] = value
            
            # Only add rows that have at least one value
            if any(row_dict.values()):
                sheet_data.append(row_dict)
        
        data[sheet_name] = sheet_data
        print(f"  → {len(sheet_data)} rows processed")
    
    # Save to JSON
    print(f"\nSaving JSON to: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n✅ Successfully converted {len(data)} sheets to JSON!")
    print("\nSummary:")
    for sheet, content in data.items():
        print(f"  {sheet}: {len(content)} rows")

def main():
    """Main entry point."""
    # Define paths
    excel_file = Path(__file__).parent / "S4LeagueAutomatisierung.xlsx"
    json_file = Path(__file__).parent / "client" / "public" / "data.json"
    
    # Check if Excel file exists
    if not excel_file.exists():
        print(f"❌ Error: Excel file not found at {excel_file}")
        print("Please place S4LeagueAutomatisierung.xlsx in the project root directory.")
        sys.exit(1)
    
    # Create output directory if it doesn't exist
    json_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Convert
    try:
        convert_excel_to_json(str(excel_file), str(json_file))
    except Exception as e:
        print(f"\n❌ Error during conversion: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
